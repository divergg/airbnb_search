import openpyxl


def send_data_to_xl(data: dict, sheet_name: str):
    workbook = openpyxl.Workbook()
    sheet = workbook.create_sheet(sheet_name)

    for i, url in enumerate(data):
        # write the URL to the first column
        sheet.cell(row=i + 1, column=1, value=url)
        # write the values of the parameters to the next columns
        j = 0
        for value in data[url]:
            sheet.cell(row=i + 1, column=j + 2, value=value)
            j += 1

    # save the workbook to a file
    workbook.save("urls.xlsx")
